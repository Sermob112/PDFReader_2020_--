
import re
import os
import openpyxl
from openpyxl.styles import Alignment

# with open("Processed_text\SS 2009.txt", "r", encoding="utf-8") as infile, open("output.txt", "w", encoding="utf-8") as outfile:
#     capture = False  # Флаг для записи нужных строк
    
#     for line in infile:
#         if line.strip().startswith(("Vessel’s", "IMO", "Total number","Flag",
#                                     "Length","scantling")):
#             outfile.write(line)




# import re

# def extract_technical_particulars_with_keywords(input_file, output_file):
#     with open(input_file, 'r', encoding='utf-8') as infile:
#         data = infile.readlines()  # Читаем файл построчно
    
#     # Регулярное выражение для поиска блока между "TECHNICAL PARTICULARS" и "Boilers"
#     start_pattern = re.compile(r'^TECHNICAL PARTICULARS\b', re.MULTILINE)
#     end_pattern = re.compile(r'^\s*Boilers\b', re.MULTILINE)

#     extracting = False  # Флаг, указывающий, что мы внутри нужного блока
#     extracted_lines = []  # Список для хранения результата

#     # Ключевые слова, которые должны остаться
#     keywords = ("Vessel’s", "IMO", "Total number", "Flag", "Length", "scantling","Contract date","Launch","Delivery date")

#     for line in data:
#         stripped_line = line.strip()

#         # Если встречаем начало блока
#         if start_pattern.match(stripped_line):
#             extracting = True

#         # Если встречаем конец блока
#         if extracting and end_pattern.match(stripped_line):
#             extracting = False

#         # Если мы находимся внутри нужного блока или строка начинается с ключевого слова
#         if extracting or stripped_line.startswith(keywords):
#             extracted_lines.append(stripped_line)

#     # Записываем результат в файл
#     if extracted_lines:
#         with open(output_file, 'w', encoding='utf-8') as outfile:
#             outfile.write("\n".join(extracted_lines) + "\n")
#         print(f"Данные записаны в '{output_file}', всего строк: {len(extracted_lines)}.")
#     else:
#         print("Не найдено нужных данных.")

# # Пример использования
# input_filename = r'Processed_text\SS 2009.txt'  # Имя входного файла
# output_filename = 'output.txt'  # Имя выходного файла
# extract_technical_particulars_with_keywords(input_filename, output_filename)

#!!!!!!!СНАЧАЛО ВСЕ В СТРОКИ!!!!!!!!!!

# with open("Processed_text\SS 2009.txt", "r", encoding="utf-8") as infile, open("output_Prod.txt", "w", encoding="utf-8") as outfile:
#     for line in infile:
#         if ":" in line:  # Проверяем, есть ли двоеточие в строке
#             outfile.write(line)  # Записываем строку в новый файл
#             if line.strip().startswith("Delivery date"):  # Если строка начинается с "Delivery date"
#                 outfile.write("-" * 40 + "\n")  # Добавляем разделитель

# print("Фильтрация завершена. Данные сохранены в output.txt")


def process_files(input_dir, output_file):
    """Обрабатывает текстовые файлы из input_dir и записывает в output_file."""

    with open(output_file, "w", encoding="utf-8") as outfile:
        for filename in os.listdir(input_dir):
            if filename.endswith(".txt"):
                filepath = os.path.join(input_dir, filename)
                with open(filepath, "r", encoding="utf-8") as infile:
                    previous_line = None  # Переменная для хранения предыдущей строки
                    for line in infile:
                        if line.strip().startswith("Shipbuilder"): # Проверяем, начинается ли текущая строка с "Shipbuilder"
                            outfile.write("-" * 40 + "\n")  # Если да, сначала пишем разделитель
                        if ":" in line:
                            outfile.write(line) # Затем записываем строку, если она содержит двоеточие
                        previous_line = line #

input_directory = "Processed_text"  # Укажите вашу папку здесь
output_filename = "output_Prod.txt"
# process_files(input_directory, output_filename)

mapping = {
    "Shipbuilder": "Shipbuilder",
    "Vessel’s name": "Vessel’s name",
    "Hull No": "Hull No",
    "Owner/Operator": "Owner/Operator",
    "Designer": "Designer",
    "Model test establishment used": "Model test establishment used",
    "Flag": "Flag",
    "IMO number": "IMO number",
    "Total number of sister ships still on order": "Total number of sister ships still on order",
    "Length oa": "Length oa",
    "Length bp": "Length bp",
    "Breadth, moulded": "Breadth moulded",
    "Depth, moulded to upper deck": "Depth moulded to upper deck",
    "Width of double skin": "Width of double skin", # нужно смотреть в контексте
    "side": "Width of double skin side",
    "bottom": "Width of double skin bottom",
    "Draught": "Draught",
    "scantling": "Draught scantling",
    "design": "Draught design",
    "Gross": "Gross",
    "Design": "Design (dwt)",
    "scantling": "scantling (dwt)",
    "Liquid volume": "Liquid volume",
    "Bunkers": "Bunkers",
    "Water ballast (m3)": "Bunkers Water ballast",
    "Daily fuel consumption": "Daily fuel consumption",
    "Main engine only": "Daily fuel consumption Main engine only",
    "Auxiliaries": "Daily fuel consumption Auxiliaries",
    "Classification society and notations": "Classification society and notations",
    "Design": "Main engine Design",
    "Model": "Main engine Model",
    "Manufacturer": "Main engine Manufacturer",
    "Number": "Main engine Number", # Вот тут нужно быть внимательным, так как Number встречается много раз
    "Type of fuel": "Main engine Type of fuel",
    "Output of each engine": "Main engine Output/speed",
    "Material": "Propeller Material",
    "Designer/Manufacturer": "Propeller Designer/Manufacturer",
    # "Number": "Propeller Number",  # Конфликт с Main engine Number
    "Pitch": "Propeller Pitch",
    "Diameter": "Propeller Diameter",
    "Speed": "Propeller Speed",
    # "Number": "Main generator engine Number", # Конфликт
    "Engine make/type": "Main generator engine Make/type",
    "Engine type": "Main generator engine Type",
    # "Type of fuel": "Main generator engine Type of fuel", # Конфликт
    "Output/speed": "Main generator engine Output/speed", # Встречается и у Main engine, нужно проверять контекст
    "Alternator make/type": "Alternator Make/type",
    "Output/speed of each set": "Alternator Output/speed",
    # "Number": "Boilers Number", # Конфликт
    "Type": "Boilers Type", # Конфликт, но можно разрешить по контексту
    "Make": "Boilers Make", # Конфликт, но можно разрешить по контексту
    "Output, each boiler": "Boilers Output",
    # "Number": "Mooring equipment Number", # Конфликт
    "Make": "Mooring equipment Make", # Конфликт, но можно разрешить по контексту
    "Type": "Mooring equipment Type", # Конфликт, но можно разрешить по контексту
    "Number of each and capacity": "Lifesaving equipment Number and capacity",
    "Make": "Lifesaving equipment Make", # Конфликт, но можно разрешить по контексту
    "Type": "Lifesaving equipment Type", # Конфликт, но можно разрешить по контексту
    # "Number": "Cargo holds Number", # Конфликт
    "Grades of cargo carried": "Grades of cargo carried",
    # "Number": "Cargo gear Number", # Конфликт
    "Type": "Cargo gear Type", # Конфликт, но можно разрешить по контексту
    "Make": "Cargo gear Make", # Конфликт, но можно разрешить по контексту
    "Stainless steel": "Cargo gear Stainless steel",
    "Capacity (each)": "Cargo gear Capacity",
    # "Make": "Fire detection system Make", # Конфликт
    # "Type": "Fire detection system Type", # Конфликт
    "Officers": "Officers",
    "Crew": "Crew",
    "Suez Repair Crew": "Suez Repair Crew",
    # "Make": "Steering gear Make", # Конфликт, неявное
    "Type": "Steering gear Type",# Конфликт, неявное
    "One-man operation": "One-man operation",
    "Make": "Fire detection system Make", # определили по предыдущему
    "Type": "Fire detection system Type", # определили по предыдущему
    "Cargo holds": "Cargo holds",
    "Engine room": "Engine room fire fighting system",
    "Local fire fighting system": "Local fire fighting system",
    "Number": "Radars Number",
    "Make": "Radars Make",
    "Models": "Radars Models",
    "Incinerator": "Incinerator",
    "Sewage plant": "Sewage plant",
    "Contract date": "Contract date",
    "Launch/float-out date": "Launch/float-out date",
    "Delivery date": "Delivery date",
}

def process_ship_data(filepath, mapping):
    """
    Читает данные из файла, добавляет заголовки на основе маппинга и контекста.

    Args:
        filepath: Путь к файлу с данными.
        mapping: Словарь-маппинг.

    Returns:
        Строка с обработанными данными.
    """
    processed_data = ""
    last_key = ""  # Для отслеживания контекста (предыдущий ключ)
    with open(filepath, 'r', encoding='utf-8') as f:  # Изменено: добавлена кодировка
        for line in f:
            line = line.strip()
            if not line:
                continue

            if line == '----------------------------------------':  # Добавлено: проверка на разделитель
                processed_data += line + "\n"
                last_key = ""  # Сбрасываем last_key
                continue

            if ":" in line:
                key, value = line.split(":", 1)
                key = key.strip()
                value = value.strip()

                # Разрешение конфликтов на основе контекста
                if key == "Number":
                    if last_key in ["Manufacturer", "Output of each engine"]:
                        mapped_key = "Main engine Number"
                    elif last_key in ["Designer/Manufacturer", "Diameter"]:
                        mapped_key = "Propeller Number"
                    elif last_key in ["Engine make/type", "Engine type", "Output/speed"]:
                        mapped_key = "Main generator engine Number"
                    elif last_key in ["Type", "Make", "Output, each boiler"]:
                        mapped_key = "Boilers Number"
                    elif last_key in ["Make", "Type"]:
                        if "sets" in value:
                            mapped_key = "Mooring equipment Number"
                        else:
                            mapped_key = "Cargo gear Number"
                    elif last_key == "Local fire fighting system":
                        mapped_key = "Radars Number"
                    else:
                        mapped_key = mapping.get(key, key) # Если не нашли в контексте, используем как есть

                elif key == "Make":
                    if last_key in ["Type", "Number", "Output, each boiler"]:
                        mapped_key = "Boilers Make"
                    elif last_key in ["Type", "Number of each and capacity"]:
                        mapped_key = "Lifesaving equipment Make"
                    elif last_key in ["Type", "Number"]:
                        if "ballast" in last_key or "fuel" in last_key:
                            mapped_key = "Mooring equipment Make"
                        elif "deepwell" in value:
                            mapped_key = "Cargo gear Make"
                        else:
                            mapped_key = "Mooring equipment Make"
                    elif last_key in ["Type", "One-man operation"]:
                        mapped_key = "Fire detection system Make"
                    elif last_key in ["Models"]:
                        mapped_key = "Radars Make"
                    else:
                        mapped_key = mapping.get(key, key)
                elif key == "Type":
                    if last_key in ["Number", "Make", "Output, each boiler"]:
                        mapped_key = "Boilers Type"
                    elif last_key in ["Number", "Make"]:
                        if "lifeboat" in value.lower():
                           mapped_key = "Lifesaving equipment Type"
                        else:
                            mapped_key = "Mooring equipment Type"
                    elif last_key in ["Make", "Capacity (each)"]:
                        mapped_key = "Cargo gear Type"
                    elif last_key in ["Make"]:
                        mapped_key = "Fire detection system Type"
                    else:
                        mapped_key = mapping.get(key, key)
                elif key == 'Output/speed' and last_key == 'Alternator make/type':
                     mapped_key = "Alternator Output/speed"
                else:
                    mapped_key = mapping.get(key, key)  # Если не нашли в маппинге, оставляем как есть

                processed_data += f"{mapped_key}: {value}\n"
                last_key = key
            else:
                # Если строка без ":", то это продолжение предыдущего значения.
                # Нужно добавить к предыдущей строке.
                processed_data = processed_data.rstrip('\n') + " " + line + "\n"
    return processed_data

# # Пример использования:
filepath = "output_Prod.txt"  # Замените на имя вашего файла
# processed_data = process_ship_data(filepath, mapping)
# # print(processed_data)

# # # Сохраняем в новый файл:
# with open("processed_ship_data.txt", "w", encoding='utf-8') as f:
#     f.write(processed_data)

import re

import openpyxl
from openpyxl.styles import Alignment

# def extract_data_from_text(filepath, mapping):
#     """
#     Извлекает данные из обработанного текста и возвращает список словарей.
#     """
#     processed_text = process_ship_data(filepath, mapping)
#     data_dict = {}
#     for line in processed_text.split('\n'):
#         if ':' in line:
#             key, value = line.split(':', 1)
#             data_dict[key.strip()] = value.strip()
#     return [data_dict]  # Возвращаем список, содержащий один словарь

# Предполагаемые заголовки для Excel (нужно обновить список)
headers = [
 "Shipbuilder", "Vessel’s name", "Hull No", "Owner/Operator", "Designer",
    "Flag", "IMO number", "Length oa", "Length bp", "Breadth moulded",
    "Depth moulded to upper deck", "Draught scantling", "Draught design",
    "Gross", "Design (dwt)", "scantling (dwt)", "Daily fuel consumption Main engine only",
    "Classification society and notations", "Main engine Design",
    "Main engine Model", "Main engine Manufacturer", "Main engine Number",
    "Main engine Type of fuel", "Main engine Output/speed",
    "Propeller Material", "Propeller Designer/Manufacturer",
    "Propeller Number", "Propeller Pitch", "Propeller Diameter",
    "Propeller Speed", "Main generator engine Number",
    "Main generator engine Make/type", "Main generator engine Type of fuel",
    "Main generator engine Output/speed", "Alternator Make/type",
    "Alternator Output/speed", "Boilers Number", "Boilers Type",
    "Boilers Make", "Boilers Output", "Mooring equipment Number",
    "Mooring equipment Type", "Mooring equipment Make",
    "Lifesaving equipment Number and capacity",
    "Lifesaving equipment Make", "Lifesaving equipment Type",
    "Cargo gear Type", "Cargo gear Make",
    "Cargo gear Stainless steel", "Cargo gear Capacity",
    "Fire detection system Make", "Fire detection system Type",
    "Cargo holds", "Engine room fire fighting system",
    "Local fire fighting system", "Radars Number", "Radars Make",
    "Radars Models", "Incinerator", "Sewage plant", "Contract date",
    "Launch/float-out date", "Delivery date"
]
def save_to_excel(data_list, headers, output_filepath):
    """
    Сохраняет список данных в Excel-файл.

    Args:
        data_list: Список словарей, где каждый словарь - строка данных.
        headers: Список заголовков столбцов.
        output_filepath: Путь к выходному Excel-файлу.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Записываем данные
    for row_num, data_row in enumerate(data_list, 2):
        for col_num, header in enumerate(headers, 1):
            value = data_row.get(header, "")
            sheet.cell(row=row_num, column=col_num, value=value)

    # Автоматическая ширина столбцов
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(output_filepath)
# Использование:
    
def extract_data_from_text(filepath, mapping):
    """
    Извлекает данные из файла с обработанным текстом и возвращает список словарей.
    """
    data_list = []
    data_dict = {}
    with open(filepath, 'r', encoding='utf-8') as f:  # Добавлено: чтение из файла
        for line in f:
            line = line.strip()
            if line == '----------------------------------------':
                if data_dict:  # Проверка, чтобы не добавлять пустой словарь в начале
                    data_list.append(data_dict)
                data_dict = {}
            elif ':' in line:
                key, value = line.split(':', 1)
                mapped_key = mapping.get(key.strip(), key.strip())  # Используем mapping
                data_dict[mapped_key] = value.strip()
    if data_dict:  # Добавляем последний словарь, если он не пустой
        data_list.append(data_dict)
    return data_list

processed_filepath = "processed_ship_data.txt"
# filepath = "processed_ship_data.txt"  # Путь к вашему файлу
output_excel_filepath = "ship_data.xlsx"  # Путь для сохранения Excel-файла

data_list = extract_data_from_text(processed_filepath, mapping)
save_to_excel(data_list, headers, output_excel_filepath)

print(f"Данные сохранены в {output_excel_filepath}")