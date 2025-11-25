# import openpyxl
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from openpyxl.worksheet.worksheet import Worksheet

# # Читаем текстовый файл
# with open("text.txt", "r", encoding="utf-8") as file:
#     lines = file.readlines()

# # Создаем новый Excel-файл
# wb = Workbook()
# ws: Worksheet = wb.active
# ws.title = "Technical Details"

# # Устанавливаем заголовки
# ws.append(["Поле", "Значение"])

# current_section = None  # Текущий раздел
# row_index = 2  # Начинаем с 2-й строки (после заголовков)

# for line in lines:
#     line = line.strip()  # Убираем пробелы и переносы строк
#     if not line:  # Пропускаем пустые строки
#         continue

#     if ":" in line:  # Это строка с данными (Поле: Значение)
#         field, value = map(str.strip, line.split(":", 1))  # Разделяем по двоеточию
#         ws.append([field, value])  # Добавляем строку
#         row_index += 1
#     else:  # Это название раздела
#         current_section = line
#         ws.append([current_section, ""])  # Записываем название раздела
#         ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)  # Объединяем колонки
#         ws[row_index][0].alignment = Alignment(horizontal="center", vertical="center")  # Центрируем текст
#         row_index += 1

# # Автоширина колонок
# for col in ws.columns:
#     max_length = 0
#     col_letter = col[0].column_letter  # Определяем букву колонки
#     for cell in col:
#         if cell.value:
#             max_length = max(max_length, len(str(cell.value)))
#     ws.column_dimensions[col_letter].width = max_length + 2  # Немного увеличиваем ширину

# # Сохраняем в Excel-файл
# output_file = "technical_details.xlsx"
# wb.save(output_file)

# print(f"Данные успешно сохранены в {output_file}")


# import re
# import pandas as pd

# # Читаем текстовый файл
# with open("formatted_text.txt", "r", encoding="utf-8") as file:
#     lines = file.readlines()

# data = []
# current_header = None

# for line in lines:
#     line = line.strip()
#     if not line:
#         continue  # Пропускаем пустые строки

#     if ":" in line:
#         # Это поле со значением
#         field, value = map(str.strip, line.split(":", 1))
#         data.append([field, value])
#     else:
#         # Это заголовок, который занимает обе колонки
#         current_header = line
#         data.append([current_header, ""])  # Заполняем заголовок в обе колонки

# # Создаем DataFrame
# df = pd.DataFrame(data, columns=["Field", "Value"])

# # Сохраняем в Excel
# output_file = "output 22.xlsx"
# df.to_excel(output_file, index=False, sheet_name="Data")

# print(f"Файл сохранен как {output_file}")


import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


input_folder = "processed_text_ss_test"
output_folder = "Results_excel"


os.makedirs(output_folder, exist_ok=True)


txt_files = [f for f in os.listdir(input_folder) if f.endswith('.txt')]

for txt_file in txt_files:
    input_file_path = os.path.join(input_folder, txt_file)
    
   
    with open(input_file_path, "r", encoding="utf-8") as file:
        lines = file.readlines()

    data = []
    current_header = None

    for line in lines:
        line = line.strip()
        if not line:
            continue  

        if ":" in line:
          
            field, value = map(str.strip, line.split(":", 1))
            data.append([field, value])

            if field == "Delivery date":
                data.append(["", ""])
                data.append(["", ""])
        else:
            # Это заголовок
            current_header = line
            data.append([current_header, ""])  


    output_file = os.path.join(output_folder, f"{os.path.splitext(txt_file)[0]}.xlsx")
    df = pd.DataFrame(data, columns=["Field", "Value"])
    df.to_excel(output_file, index=False, sheet_name="Data")

   
    wb = load_workbook(output_file)
    ws = wb.active

   
    for row in range(2, ws.max_row + 1):  
        field_cell = ws[f"A{row}"]
        value_cell = ws[f"B{row}"]

        if field_cell.value and not value_cell.value: 
            field_cell.font = Font(bold=True)  
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)  


    wb.save(output_file)

    print(f"Файл '{txt_file}' сохранен как '{output_file}'")
