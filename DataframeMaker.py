# import re
# import pandas as pd

# # Читаем данные из файла
# with open("output_Prod.txt", "r", encoding="utf-8") as file:
#     data = file.read()

# # Разделяем по границам записей
# ship_entries = data.split('----------------------------------------')

# # Поля, которые нас интересуют
# fields = ["Shipbuilder", "Vessel’s name", "Hull No"]

# # Список для хранения данных
# ship_data = []

# # Обрабатываем каждую запись
# for entry in ship_entries:
#     ship_info = {}
    
#     # Ищем строки вида "Ключ: Значение"
#     for line in entry.split("\n"):
#         match = re.match(r"(.+?):\s*(.+)", line)
#         if match:
#             key, value = match.groups()
#             key = key.strip()
#             value = value.strip()
            
#             # Если ключ важный, сохраняем
#             if key in fields:
#                 ship_info[key] = value
    
#     # Добавляем в таблицу, если есть нужные поля
#     if ship_info:
#         ship_data.append([
#             ship_info.get("Shipbuilder", "N/A"),
#             ship_info.get("Vessel’s name", "N/A"),
#             ship_info.get("Hull No", "N/A")
#         ])

# # Создаем DataFrame
# df = pd.DataFrame(ship_data, columns=["Shipbuilder", "Vessel’s name", "Hull No"])

# # Добавляем нумерацию
# df.index += 1
# df.index.name = "№"

# # Выводим таблицу
# print(df)

# # Экспорт в CSV (если нужно)
# df.to_csv("ships_summary.csv", encoding="utf-8", index=True)






import openpyxl
from openpyxl.styles import Alignment

# ... (предыдущий код с функцией process_ship_data и словарем mapping) ...

def save_to_excel(data_list, headers, output_filepath):
    """
    Сохраняет список данных в Excel-файл.

    Args:
        data_list: Список словарей, где каждый словарь - строка данных.
        headers: Список заголовков столбцов.
        output_filepath: Путь к выходному Excel-файлу.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Записываем данные
    for row_num, data_row in enumerate(data_list, 2):
        for col_num, header in enumerate(headers, 1):
            value = data_row.get(header, "")
            sheet.cell(row=row_num, column=col_num, value=value)

    # Автоматическая ширина столбцов
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(output_filepath)

def extract_data_from_text(filepath, mapping):
    """
    Извлекает данные из обработанного текста и возвращает список словарей.
    """
    processed_text = process_ship_data(filepath, mapping)
    data_list = []  # Изменено: создаем список для хранения данных
    data_dict = {}
    for line in processed_text.split('\n'):
        if line.strip() == '----------------------------------------':  # Добавлено: разделитель между судами
            data_list.append(data_dict)
            data_dict = {}
        elif ':' in line:
            key, value = line.split(':', 1)
            data_dict[key.strip()] = value.strip()
    data_list.append(data_dict)  # Добавлено: добавляем последний словарь после цикла
    return data_list

# ... (словарь mapping) ...

# Предполагаемые заголовки для Excel (нужно обновить список)
headers = [
    "Shipbuilder", "Vessel’s name", "Hull No", "Owner/Operator", "Designer",
    "Flag", "IMO number", "Length oa", "Length bp", "Breadth moulded",
    "Depth moulded to upper deck", "Draught scantling", "Draught design",
    "Gross", "Design (dwt)", "scantling (dwt)", "Daily fuel consumption Main engine only",
    "Classification society and notations", "Main engine Design",
    "Main engine Model", "Main engine Manufacturer", "Main engine Number",
    "Main engine Type of fuel", "Main engine Output/speed",
    "Propeller Material", "Propeller Designer/Manufacturer",
    "Propeller Number", "Propeller Pitch", "Propeller Diameter",
    "Propeller Speed", "Main generator engine Number",
    "Main generator engine Make/type", "Main generator engine Type of fuel",
    "Main generator engine Output/speed", "Alternator Make/type",
    "Alternator Output/speed", "Boilers Number", "Boilers Type",
    "Boilers Make", "Boilers Output", "Mooring equipment Number",
    "Mooring equipment Type", "Mooring equipment Make",
    "Lifesaving equipment Number and capacity",
    "Lifesaving equipment Make", "Lifesaving equipment Type",
    "Cargo gear Type", "Cargo gear Make",
    "Cargo gear Stainless steel", "Cargo gear Capacity",
    "Fire detection system Make", "Fire detection system Type",
    "Cargo holds", "Engine room fire fighting system",
    "Local fire fighting system", "Radars Number", "Radars Make",
    "Radars Models", "Incinerator", "Sewage plant", "Contract date",
    "Launch/float-out date", "Delivery date"
]

# Использование:
filepath = "ship_data.txt"  # Путь к вашему файлу (теперь содержит несколько судов, разделенных '---')
output_excel_filepath = "ship_data.xlsx"  # Путь для сохранения Excel-файла

data_list = extract_data_from_text(filepath, mapping)
save_to_excel(data_list, headers, output_excel_filepath)

print(f"Данные сохранены в {output_excel_filepath}")